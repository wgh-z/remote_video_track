#codeing=utf-8
import pandas as pd
import openpyxl as op
import os
import sys
base_path=os.getcwd()
sys.path.append(base_path)
# open_excel=openpyxl.load_workbook(base_path+"/Case/android4.0.1.xlsx")
# sheet_name=open_excel.sheetnames
# excel_value=open_excel[sheet_name[1]]
# print (excel_value)
# print (excel_value.cell(1,3).value)
# print (excel_value.max_row)


def read(path, column=None, ignore=True):
    """
    :param path:要读取的文件路径
    :param column:要读取的文件列
    :param ignore:是否忽略空值,True为忽略
    :return:读取到的值列表
    """
    data_frame = pd.read_excel(path, usecols=column, header=0, keep_default_na=False)
    _list = data_frame.values.tolist()
    if ignore:
        while [''] in _list:
            _list.remove([''])
    return _list

def write(path, data, row, column=5):
    bg = op.load_workbook(path)  # 应先将excel文件放入到工作目录下
    sheet = bg["Sheet1"]         # “Sheet1”表示将数据写入到excel文件的sheet1下
    sheet.cell(row, column, str(data))  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
    bg.save(path)  # 对文件进行保存


def clear(path):
    wb = op.load_workbook(path)
    wr = wb.active
    row = wb[wb.sheetnames[0]].max_row
    column = wb[wb.sheetnames[0]].max_column
    for i in range(2,row+1):
        for j in range(1,column+1):
            wr.cell(row=i,column=j,value="")
            # print(self.get_sheet_data().cell(row=i,column=j,value=""))
            wb.save(path)

class HandExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        open_excel=op.load_workbook(base_path+"/Case/case.xlsx")
        return open_excel
    def get_sheet_data(self,index=None):
        '''
        加载所有sheet的内容
        '''
        sheet_name = self.load_excel().sheetnames
        if index == None:
 
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data
    
    def get_cell_value(self,row,col):
        '''
        获取某一个单元格的内容
    
        '''
        data = self.get_sheet_data().cell(row=row,column=col).value
        return data
    
    def get_rows(self):
        '''
        获取行数
        '''
        row = self.get_sheet_data().max_row
        return row
 
    def get_column(self):
        '''
        获取列数
        '''
        column = self.get_sheet_data().max_column
        return column
 
    def get_rows_value(self,row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_sheet_data()[row]:
 
            row_list.append(i.value)
        return row_list
    
    def excel_write_data(self,row,cols,data):
        '''
        往excel表格里写入数据
        '''
        wb=self.load_excel()
        '''
        激活excel表格
        '''
        wr=wb.active
        wr.cell(row,cols,data)
        # self.get_sheet_data().cell(row,cols,data)
        wb.save(base_path+"/Case/case.xlsx")
 
    def excel_data_clear(self):
        '''
        清除excel表格中从第2行开始的数据
        '''
        wb=self.load_excel()
        '''
        激活excel表格
        '''
        wr=wb.active
 
        
        row=self.get_rows()
        column=self.get_column()
        for i in range(2,row+1):
            for j in range(1,column+1):
                wr.cell(row=i,column=j,value="")
                # print(self.get_sheet_data().cell(row=i,column=j,value=""))
                wb.save(base_path+"/Case/case.xlsx")
                
 
 
        
    def get_colsa_value(self,col):
        '''
        获取Excel中某一列的数据转化为列表的形式存储
        '''
        #k=dict(self.get_sheet_data().iter_cols(10,10))
        col_list=[]
        cols=self.get_sheet_data().iter_cols(col,col)
 
        for col in cols:
            #print(col)
            for cell in col:
                #print(cell.value)
                col_list.append(cell.value)
        return col_list
 
            
        
        
 
 
handle_excel = HandExcel()
if __name__ == "__main__":
    # handle_excel = HandExcel()
    # print(base_path)
    # print(handle_excel.load_excel())
    # print('===========================================================')
    # print(handle_excel.get_sheet_data(1))
    # print('===========================================================')
    # print(handle_excel.get_cell_value(3,2))
    #print('===========================================================')
    #print(handle_excel.get_rows_value(2))
    # print(handle_excel.get_rows())
    # print(handle_excel.get_column())
    # print(handle_excel.excel_data_clear())
    # handle_excel.excel_write_data(2,9,"通过")
    #print(handle_excel.get_colsa_value(10))

    write(r"测试.xlsx", [1,2], 1, 1)
    write(r"测试.xlsx", [1,2,3], 2, 1)